import sys
import pprint
import openpyxl
import textwrap

from TSHTDM.DataNode import DataNode

from TSHTDM.MDNode import MDNode
from TSHTDM.DataNode1D import DataNode1D
from TSHTDM.DataNode2D import DataNode2D
from argparse import Action
from TSHTDM.SyntaxRec import SyntaxRec

from TSHExcel.TSHpandas import TSHpandas
from TSHExcel.TSHopenpyxl import TSHopenpyxl
from TSHGlobal import constants
from TSHFS.TSHFile import TSHFile
from TSH.TSHKeyMod import TSHKeyMod
from TSHGlobal import constants as c
from TSHDict import TSHDict
from TSHComment.TSHComment import TSHComment

class RecurseMD():
    def PrintDebugInfo(self, debugId):
        pass
    
    def __init__(self):
        self.keyMod = None
        
        self.udFile = ""
        self.mdFile = ""
        
        self.udKey = ""        
        self.mdKey = ""
        
        self.udkType = ""
        self.mdkType = ""
        
        self.sheetName = ""
        self.debugId = 1
        #self.wbMD = None
        self.topMDObj = None
        self.topUDObj = None
        
#        self.MwareKeyAss = {}
#        self.DwareKeyAss = {}
        
        self.TwareComment = None
        self.DwareComment = None
        self.MwareComment = None    
        self.THarnComment = None
        self.funcAddr     = None
        
        
        self.MDKDelimiter = None
        self.UDKDelimiter = None

    @property
    def funcAdrr(self):
        return self.__funcAddr
    
    @funcAdrr.setter
    def funcAdrr(self, funcAddr):
        self.__funcAddr = funcAddr
        
    @property
    def THarnComment(self):
        return self.__THarnComment
    
    @THarnComment.setter
    def THarnComment(self, THarnComment):
        self.__THarnComment = THarnComment

    @property
    def TwareComment(self):
        return self.__TwareComment
    
    @TwareComment.setter
    def TwareComment(self, TwareComment):
        self.__TwareComment = TwareComment
    
    @property
    def DwareComment(self):
        return self.__DwareComment
    
    @DwareComment.setter
    def DwareComment(self, DwareComment):
        self.__DwareComment = DwareComment
        
    @property
    def MwareComment(self):
        return self.__MwareComment
    
    @MwareComment.setter
    def MwareComment(self, MwareComment):
        self.__MwareComment = MwareComment
        
    @property
    def udFile(self):
        return self.__udFile
    
    @udFile.setter
    def udFile(self, value):
        self.__udFile = value
    
    @property
    def mdFile(self):
        return self.__mdFile
    
    @mdFile.setter
    def mdFile(self, value):
        self.__mdFile = value
    
    @property
    def sheetName(self):
        return self.__sheetName
    
    @sheetName.setter
    def sheetName(self, value):
        self.__sheetName = value
    
    @property
    def mdKey(self):
        return self.__mdKey
    
    @mdKey.setter
    def mdKey(self, value):
        self.__mdKey = value
        
    @property
    def udKey(self):
        return self.__udKey
    
    @udKey.setter
    def udKey(self, value):
        self.__udKey = value

    @property
    def mdkType(self):
        return self.__mdkType
    
    @mdkType.setter
    def mdkType(self, value):
        self.__mdkType = value
    
    @property
    def udkType(self):
        return self.__udkType
    
    @udkType.setter
    def udkType(self, value):
        self.__udkType = value
        
    #------------------------------------------------------------------------------------------- 
    def ReadComment(self, mdFile):
        workbook = TSHopenpyxl.load_workbook(mdFile)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment:
                    print(cell.comment.text)    
                    
    #------------------------------------------------------------------------------------------- 
    #------------------------------------------------------------------------------------------- 
    
    def GetPadStr(self, indentLvl):
        i = 0
        padStr = ""
        for i in range(indentLvl):
            padStr = padStr + '  '
        return padStr
    
    #------------------------------------------------------------------------------------------- 
    #------------------------------------------------------------------------------------------- 
    
    def TraverseMD(self, MDList0, indentLvl):
        Ind = 0
        while(Ind < MDList0.GetDataNode1DLen()):
            DataNode0 = MDList0.GetDataNode(Ind)
            print(DataNode0.DispObjVal(self.GetPadStr(indentLvl-1)))
            if ("array" == DataNode0.GetDataType()):
                self.TraverseMD(DataNode0.GetDataNode1D(), indentLvl+1)
            Ind = Ind + 1
                
    #------------------------------------------------------------------------------------------- 
    #------------------------------------------------------------------------------------------- 
    def GetMD(self, DataList0, wbObj, sheetName, row, col, indentLvl, MoreColumns, ignoreRecursion):
        #print("GetMD:" + sheetName)            
        MwareDatAss = {}
        MwareKeyAss = {}
        
        sheetObj = self.topUDObj.get_sheetObj(sheetName)
        if (None != sheetObj):
            pass
        else:
            return -1

        keyMod2 = TSHKeyMod(sheetObj, self.MwareComment, self.MDKDelimiter, self.UDKDelimiter)
        MwareKeyAss = keyMod2.GetMDKeys(self.mdKey, self.mdkType)
        col = keyMod2.EOMDColInd + 1
        #for key in (MwareKeyAss):
        #    print("MwareKeyAss[" + str(key) + "]:" + str(MwareKeyAss[key]))
            #MwareDatAss[key] = sheetObj.cell(DataTypeCol, col).value
        
        
        
        AttributeNameAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "AttributeName" 
        DataTypeAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "DataType"
        TabAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "Tab"
        
        
        AttributeNameCol = -1
        DataTypeCol = -1
        TabCol = -1
        for key in (MwareKeyAss):
            #print(str(key) + ":" + str(MwareKeyAss[key]))
            val = MwareKeyAss[key]
            
            x = val.rfind(AttributeNameAttr, 0, len(val))
            if (-1 != x):
                AttributeNameCol = int(key)

            x = val.rfind(DataTypeAttr, 0, len(val))
            if (-1 != x):
                DataTypeCol = int(key)
                    
            x = val.rfind(TabAttr, 0, len(val))
            if (-1 != x):
                TabCol = int(key)

        if (-1 == AttributeNameCol):
            print("Error: There is a problem with the MD Key Setup for " + str(self.mdKey))
            return -1
        
        MDNode0 = MDNode(MwareKeyAss)
        DataList0.AddDataNode(MDNode0)

        while (True == MoreColumns):

            commentCellValue = sheetObj.cell(1, col).value
            ercRet = self.MwareComment.isAComment(commentCellValue)
            ercRet = self.MwareComment.ProcessCommentStatus()
            if (ercRet == -1):
                return ercRet
            AttributeName = sheetObj.cell(AttributeNameCol, col).value
            if (ercRet == 1) and (AttributeName is None):
                col += 1
                continue
            elif (ercRet == 1) and (AttributeName is not None):
                col += 1
                continue
            elif (ercRet == 0) and (AttributeName is None):
                MoreColumns = False
            elif (ercRet == 0) and (AttributeName is not None):

                for  key in (MwareKeyAss):
                    #if (-1 != key):
                    MwareDatAss[key] = sheetObj.cell(int(key), col).value
                
                DataType = sheetObj.cell(DataTypeCol, col).value
                if (DataType is None): DataType = "None"
                
                Tab = sheetObj.cell(TabCol, col).value
                if (Tab is None): Tab = "None"
                #self.PrintMDCol(AttributeName, DataType, Tab)

                MwareDatAss["SSColInd"] = col
                MDNode0 = MDNode(MwareDatAss)
                
                if ("array" == DataType):
                    if ("Y" == ignoreRecursion):
                        DataList0.AddDataNode(MDNode0)
                    else:
                        MoreColumns = True
                        indentLvl += 1
                        DataList0.AddDataNode(MDNode0)
                        ercRet = self.GetMD(MDNode0.GetDataNode1D(), wbObj, Tab, 1, 1, indentLvl, MoreColumns, "")
                        if (-1 == ercRet):
                            return ercRet
                        indentLvl -= 1
                else:
                    DataList0.AddDataNode(MDNode0)
                col += 1
            else:
                return -1
        
        

        return 0
                
    #------------------------------------------------------------------------------------------- 
    #------------------------------------------------------------------------------------------- 

    def GetMD1(self, DataList0, wbObj, sheetName, row, col, indentLvl, MoreColumns):            
        MwareDatAss = {}
        MwareKeyAss = {}
        
        sheetObj = self.topUDObj.get_sheetObj(sheetName)
        if (None != sheetObj):
            pass
        else:
            return -1

        keyMod2 = TSHKeyMod(sheetObj, self.MwareComment, self.MDKDelimiter, self.UDKDelimiter)
        MwareKeyAss = keyMod2.GetMDKeys(self.mdKey, self.mdkType)
        col = keyMod2.EOMDColInd + 1
        #for key in (MwareKeyAss):
        #    print("MwareKeyAss[" + str(key) + "]:" + str(MwareKeyAss[key]))
            #MwareDatAss[key] = sheetObj.cell(DataTypeCol, col).value
        
        
        
        AttributeNameAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "AttributeName" 
        DataTypeAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "DataType"
        TabAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "Tab"
        
        
        AttributeNameCol = -1
        DataTypeCol = -1
        TabCol = -1
        for key in (MwareKeyAss):
            #print(str(key) + ":" + str(MwareKeyAss[key]))
            val = MwareKeyAss[key]
            
            x = val.rfind(AttributeNameAttr, 0, len(val))
            if (-1 != x):
                AttributeNameCol = int(key)

            x = val.rfind(DataTypeAttr, 0, len(val))
            if (-1 != x):
                DataTypeCol = int(key)
                    
            x = val.rfind(TabAttr, 0, len(val))
            if (-1 != x):
                TabCol = int(key)

        if (-1 == AttributeNameCol):
            print("Error: There is a problem with the MD Key Setup for " + str(self.mdKey))
            return -1
        
        MDNode0 = MDNode(MwareKeyAss)
        DataList0.AddDataNode(MDNode0)

        while (True == MoreColumns):

            commentCellValue = sheetObj.cell(1, col).value
            ercRet = self.MwareComment.isAComment(commentCellValue)
            ercRet = self.MwareComment.ProcessCommentStatus()
            if (ercRet == -1):
                return ercRet
            AttributeName = sheetObj.cell(AttributeNameCol, col).value
            if (ercRet == 1) and (AttributeName is None):
                col += 1
                continue
            elif (ercRet == 1) and (AttributeName is not None):
                col += 1
                continue
            elif (ercRet == 0) and (AttributeName is None):
                MoreColumns = False
            elif (ercRet == 0) and (AttributeName is not None):

                for  key in (MwareKeyAss):
                    #if (-1 != key):
                    MwareDatAss[key] = sheetObj.cell(int(key), col).value
                
                DataType = sheetObj.cell(DataTypeCol, col).value
                if (DataType is None): DataType = "None"
                
                Tab = sheetObj.cell(TabCol, col).value
                if (Tab is None): Tab = "None"
                #self.PrintMDCol(AttributeName, DataType, Tab)

                MwareDatAss["XLColInd"] = col
                MDNode0 = MDNode(MwareDatAss)
                
                if ("array" == DataType):
                    MoreColumns = True
                    indentLvl += 1
                    DataList0.AddDataNode(MDNode0)
                    ercRet = self.GetMD(MDNode0.GetDataNode1D(), wbObj, Tab, 1, 1, indentLvl, MoreColumns)
                    if (-1 == ercRet):
                        return ercRet
                    indentLvl -= 1
                else:
                    DataList0.AddDataNode(MDNode0)
                col += 1
            else:
                return -1
        
        

        return 0
                
    #------------------------------------------------------------------------------------------- 
    #------------------------------------------------------------------------------------------- 
    def GetUDLocation(self, wbObj, sheetObj, row, col, TCNameInd, udKey):
        self.GetUDLocWithSingKey(self, wbObj, sheetObj, row, col, TCNameInd, udKey)
        #self.keyMod.GetUDKeys(self.udKey, "Single")
        pass
        
    def GetUDLocWithMultKey(self, wbObj, sheetObj, row, col, TCNameInd, udKey):
        pass
    
    def GetUDLocWithSingKey(self, wbObj, sheetObj, row, col, TCNameInd, udKey, keyType):
        
        TCName = udKey + c.TSH_UDKDelimiter + str(f'{TCNameInd:02}')

        retVal = -1
        endOfData = -1
        skip = -2
            
        loop = True
           
        while (True == loop):
            comment = sheetObj.cell(row, 1).value
#            TestCaseDelimiter = sheetObj.cell(row, 2).value
#            if (TestCaseDelimiter is None):
#                TestCaseDelimiter = c.TSH_UDKDelimiter
            XLTCName = sheetObj.cell(row, constants.UDKHDRCol["UDK01"]).value
            TCName = udKey + c.TSH_UDKDelimiter + str(f'{TCNameInd:02}')
                        
            if (c.TSH_SingRCComment == comment):
                if (TCName == XLTCName):
                    retVal = skip
                    loop = False
                else:
                    row += 1
            else:
                if (TCName == XLTCName):
                    retVal = row
                    loop = False
                elif (XLTCName is None):
                    retVal = endOfData
                    loop = False
                else:
                    row += 1
    
        return (retVal, TCName, TCNameInd)
    
    #------------------------------------------------------------------------------------------- 
    def ProcessCellComment(self, sheetObj, XLRowInd, XLColInd):
        cell  = sheetObj.cell(XLRowInd, XLColInd)
        if (cell is not None):
            if (cell.comment is not None):
                comment = str(cell.comment)
                #xArr = comment.split()
                xArr = comment.splitlines()
                #xArr = comment.split('\n')
                for xEl in (xArr):
                    xEl = xEl.replace("Comment: ", "", 1)
                    xEl = xEl.replace(" by Alexander Raj", "", 1)
                    #print ("Line:" + xEl)
                    xArr1 = xEl.split('=')
                    print (len(xArr1))
                    if (len(xArr1) == 2):
                        xArr1[0] = xArr1[0].strip()
                        xArr1[1] = xArr1[1].strip()
                        #xArr1[0] = strip(xArr1[0])
                        #xArr1[1] = strip(xArr1[1])
                        print("key:", xArr1[0])
                        print("val:", xArr1[1])
                            #print ("-------------comment:" + str(comment))
                    
#print(sheetObj.cell(XLRowInd, XLColInd). .comment.text)
#print (":WBS:"+sheetName+":TCP:"+udKey+":TC:"+TCName+":MD:"+str(MDColInd)+":UD:"+str(UDRowInd)+":"+str(UDColInd)+":XL:"+str(XLRowInd) + ":" + str(XLColInd)+":DT:"+MDNode0.GetDataType()+":val:"+str(dataVal))        
#print (str(dataVal))

    #-------------------------------------------------------------------------------------------

    def TSHTreeNodesAdd(self, dataArr, ssLineNum):
        for dataInd in range(len(dataArr)):
            if (dataArr[dataInd] == None):
                break
        dataTot = dataInd
        
        currRootNode = self.root
        
        for dataInd in range(dataTot):
            dataEl = dataArr[dataInd]
            print(dataEl)            
            self.TSHTreeNodeAdd(currRootNode, dataEl, ssLineNum) 
    
    #------------------------------------------------------------------------------------------- 
    
    def LookupSyntaxRec(self, SyntaxList, SyntaxID):
        totCnt = SyntaxList.GetDataNode1DLen()
        ind = 0
        while (ind < totCnt):
            tSyntaxRec = SyntaxList.GetDataNode(ind)
            if (SyntaxID == tSyntaxRec.GetSyntaxID()):
                return tSyntaxRec
            else:
                ind += 1
        return None

    #-------------------------------------------------------------------------------------------

    def GetMDSyntax(self, sheetObj, SyntaxList):
        
        XLRowInd, XLColInd = self.keyMod.GetEOMDRowAndCol()
        
        if ((-1 == XLRowInd)or(-1 == XLColInd)):
            return
        else:
            XLColInd += 1
            endOfRecords = False
            while(False == endOfRecords):
                XLRowInd += 1
                
                syntaxID = sheetObj.cell(XLRowInd, XLColInd).value
                action = sheetObj.cell(XLRowInd, XLColInd+1).value
                resourceType = sheetObj.cell(XLRowInd, XLColInd+2).value
                
    
                
                if (syntaxID is not None):
                    #print("syntaxID:" + str(syntaxID))
                    #if (action is not None):
                    #    print("action:" + action)
                    #if (resourceType is not None):
                    #    print("resourceType:" + resourceType)
                    SyntaxRec01 = SyntaxRec(syntaxID, action, resourceType)
                    SyntaxList.AddDataNode(SyntaxRec01)
                else:
                    endOfRecords = True
        
        
        return 0
    #-------------------------------------------------------------------------------------------
    
    def TrimKeys(self, sheetName, keyMod, DwareKeyAss, udKey, indentLvl): 
        keyToDelArr = []
        print("udKey" + udKey)
        debug = 1
        if (1 == debug):
            print ("DwareKeyAssOriginal:" + sheetName)
            print (DwareKeyAss)
        for key in (DwareKeyAss):
            offset = DwareKeyAss[key].find(udKey + self.UDKDelimiter)
            if (0 == offset):
                DwareKeyElArr = DwareKeyAss[key].split(self.UDKDelimiter)
                if (len(DwareKeyElArr) == (indentLvl+2)):
                    print(str(key) + ":" + DwareKeyAss[key])
                    pass
                else:
                    #print(str(key) + ":" + DwareKeyAss[key])
                    debug = 1
                    keyToDelArr.append(key)
            else:
                #print(str(key) + ":" + DwareKeyAss[key])
                debug = 1
                keyToDelArr.append(key)
                #DwareKeyAss.pop(key, None)
        for keyToDelInd in range(len(keyToDelArr)):
            del (DwareKeyAss[keyToDelArr[keyToDelInd]])
        if (1 == debug):
            print ("DwareKeyAssTrimmed:" + sheetName)
            print (DwareKeyAss)
            print("-----------")
            keyMod.TSHTreeDisp(keyMod.vKTree, indentLvl)
            print("-----------")
        return DwareKeyAss

    #-------------------------------------------------------------------------------------------
    def GetUD(self,
              SyntaxList, 
              MDList0, 
              UDList0, 
              wbObj,
              sheetName, 
              TCNameInd, 
              udKey,
              UDRowInd,
              indentLvl,
              rootName):
    #-------------------------------------------------------------------------------------------
        
        padStr = ""
        for indentInd in range(indentLvl):
            padStr = padStr + " "

        sheetObj = self.topUDObj.get_sheetObj(sheetName)
        if (None != sheetObj):
            pass
        else:
            return -1
        sheetComment = TSHComment()
        sheetComment.singCommentPat = self.DwareComment.singCommentPat
        sheetComment.multCommentBegPat = self.DwareComment.multCommentBegPat
        sheetComment.multCommentEndPat = self.DwareComment.multCommentEndPat

        keyMod = TSHKeyMod(sheetObj, sheetComment, self.MDKDelimiter, self.UDKDelimiter)
        if ("M" == self.udkType):
            DwareKeyAss = keyMod.GetUDKeys(rootName, self.udkType)
        elif ("S" == self.udkType):
            DwareKeyAss = keyMod.GetUDKeys(udKey, self.udkType)
        else:
            print("Error")
            return -1
        DwareKeyAss = self.TrimKeys(sheetName, keyMod, DwareKeyAss, udKey, indentLvl)
        x = 0
        for key in (DwareKeyAss):
            x += 1
            XLRowInd = int(key)
            XLColInd = keyMod.EOMDColInd + x

            #Create data for start of a new row and add it to two dimensional array
            UD1D = DataNode1D()
            UDList0.AddDataNode1D(UD1D)
            UDRowInd += 1
            
            MDColInd = 0
            UDColInd = 0

            #Store encoded form of key = single key format
            keyAss = {}
            keyAss[key] = DwareKeyAss[key]
            
            UDNode0 = DataNode(keyAss)
            UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0)

            TCName = DwareKeyAss[key]
            
            MDNode0 = MDList0.GetDataNode(MDColInd)
            MwareKeyAss = MDNode0.MDAss
            
            MDColInd += 1

            AttributeNameAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "AttributeName" 
            DataTypeAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "DataType"
            TabAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "Tab"
        
            #AttributeNameAttr = self.mdKey + "_TSH_AttributeName"
            #DataTypeAttr = self.mdKey + "_TSH_DataType"
            #TabAttr = self.mdKey + "_TSH_Tab"
            
            AttributeNameCol = -1
            DataTypeCol = -1
            TabCol = -1
            for key in (MwareKeyAss):
                val = MwareKeyAss[key]
                
                x = val.rfind(AttributeNameAttr, 0, len(val))
                if (-1 != x):
                    AttributeNameCol = key
    
                x = val.rfind(DataTypeAttr, 0, len(val))
                if (-1 != x):
                    DataTypeCol = key
                        
                x = val.rfind(TabAttr, 0, len(val))
                if (-1 != x):
                    TabCol = key

                
            if (-1 == AttributeNameCol):
                print("Error: There is a problem with the MD Key Setup for " + str(self.mdKey))
                return -1

            #Store data vector for each MDColumn
            while(MDColInd < MDList0.GetDataNode1DLen()):
                MDNode0 = MDList0.GetDataNode(MDColInd)
                MwareDatAss = MDNode0.MDAss
                            
                XLColInd = MwareDatAss["SSColInd"]
                dataVal = sheetObj.cell(XLRowInd, XLColInd).value
                
#-------------------------------------------------------------------------
                #print("dataVal:" + str(MDColInd) + ":" + str(dataVal))
#-------------------------------------------------------------------------                
                if (None != MwareDatAss[DataTypeCol]):
                    if ("array" == MwareDatAss[DataTypeCol]):
                        if ("n" != dataVal):
                            UDNode0 = DataNode(None)
                            indexOfLastNode = UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0)
                            ercRet = self.GetUD(SyntaxList, 
                                                MDNode0.GetDataNode1D(), 
                                                UDNode0.GetDataNode2D(), 
                                                wbObj, 
                                                MwareDatAss[TabCol], 
                                                1, 
                                                TCName, 
                                                -1, 
                                                indentLvl+1,
                                                rootName)
                            childTot = UDList0.GetDataNode1D(UDRowInd).GetLastDataNode().SumChildCount()
                            #self.funcAddr(padStr, MwareKeyAss, MwareDatAss, dataVal, childTot)                            
                            pass
                        else:
                            pass
                    else:
                        UDNode0 = DataNode(dataVal)
                        newLen = UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0)
                        #self.funcAddr(padStr, MwareKeyAss, MwareDatAss, dataVal, UDNode0.SumChildCount())                            

                else:
                        UDNode0 = DataNode(dataVal)
                        newLen = UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0)
                        #self.funcAddr(padStr, MwareKeyAss, MwareDatAss, dataVal, UDNode0.SumChildCount())                            

                MDColInd += 1
                UDColInd += 1                
        return 0

    #-------------------------------------------------------------------------------------------     
    
    def TraverseUD(self, MDList0, UDList0, indentLvl):        
        padStr = ""
        for indentLvlInd in range(indentLvl):
            padStr = padStr + "       "

        dNode2D = UDList0.GetDataNode2D()
        if (None == dNode2D):
            return
        
        x = len(dNode2D)
        for dNode2DInd in range(len(dNode2D)):
            dNode1D = dNode2D[dNode2DInd]
            if (None == dNode1D):
                continue


            MDColInd = 0
            MDNode0 = MDList0.GetDataNode(MDColInd)
            MwareKeyAss = MDNode0.MDAss
            
            MDColInd += 1
            while(MDColInd < MDList0.GetDataNode1DLen()):
                #print("MDColInd" + str(MDColInd))
                MDNode0 = MDList0.GetDataNode(MDColInd)
                MwareDatAss = MDNode0.MDAss
                
                AttributeNameAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "AttributeName" 
                DataTypeAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "DataType"
                #TabAttr = self.mdKey + c.TSH_MDKDelimiter + "TSH" + c.TSH_MDKDelimiter + "Tab"

                #AttributeNameAttr = self.mdKey + "_TSH_AttributeName"
                #DataTypeAttr = self.mdKey + "_TSH_DataType"
                DataTypeCol = -1
                for key in (MwareKeyAss):
                    val = MwareKeyAss[key]                
                    x = val.rfind(AttributeNameAttr, 0, len(val))
                    if (-1 != x):
                        AttributeNameCol = key

                    x = val.rfind(DataTypeAttr, 0, len(val))
                    if (-1 != x):
                        DataTypeCol = key
                    
                    #NEED TO CHECK FOR ERROR AND BAIL OUT
        #for dNode1DInd in range(1, dNode1D.GetDataNode1DLen()):
                #XLColInd = MwareDatAss["XLColInd"]
                #dataVal = sheetObj.cell(XLRowInd, XLColInd).value
                #self.funcAddr(padStr, MwareKeyAss, MwareDatAss, dataVal)
    
                dNode = dNode1D.GetDataNode(MDColInd)
                #dNode.NoneElCount()
                #print("d1:" + str(dNode.NoneNode))
                if (None != dNode):
                    ipDict = {}
                    ipDict["padStr"]       = padStr
                    ipDict["MwareKeyAss"]  = MwareKeyAss
                    ipDict["MwareDatAss"]  = MwareDatAss 
                    ipDict["dataVal"]      = dNode.GetDataVal()
                    ipDict["childCount"]   = dNode.SumChildCount()
                    ipDict["MDKDelimiter"] = self.MDKDelimiter 
                    ipDict["UDKDelimiter"] = self.UDKDelimiter
                    opDict = {}

                    if (None != self.funcAddr):
                        self.funcAddr(ipDict, opDict)
                    
                    if (None != MwareDatAss[DataTypeCol]):
                        if ("array" == MwareDatAss[DataTypeCol]):
                            #if (0 > dNode.GetDataNode2D().GetDataNode2DLen()):
                            self.TraverseUD( MDNode0.GetDataNode1D(), 
                                                dNode.GetDataNode2D(), 
                                                indentLvl + 1)

                    

                #dataVal = dNode.GetDataVal()
                #print(padStr + str(dataVal))
                MDColInd += 1
                       
        pass
    #------------------------------------------------------------------------------------------- 
     
    def PrintMDCol(self, AttributeName, DataType, Tab):
        print("AttributeName:" + str(AttributeName))
        print("DataType:" + str(DataType))
        print("Tab:" + str(Tab))
        pass

    #------------------------------------------------------------------------------------------- 
    def mainFunc(self):
        SyntaxList = DataNode1D()
        MDList0 = DataNode1D()
        UDList0 = DataNode2D()
        
        sheetObj = None
        self.topMDObj = TSHopenpyxl(self.mdFile)
        mdWBObj = self.topMDObj.load_workbook()
        if (mdWBObj != None):
            pass
        else:
            return -2

        
        sheetObj = self.topMDObj.get_sheetObj("Syntax")
        if (None != sheetObj): 
            pass
        else:
            return -1

        self.keyMod = TSHKeyMod(sheetObj, self.THarnComment, self.MDKDelimiter, self.UDKDelimiter)
        self.GetMDSyntax(sheetObj, SyntaxList) 
        
        self.topUDObj = TSHopenpyxl(self.udFile)
        udWBObj = self.topUDObj.load_workbook()
        if (udWBObj != None):
            pass
        else:
            return -2
                        
        row = 1 #c.MDKHdrRow["MDHeader"]
        col = 1 #c.MDKHdrCol["MDK01"]
        indentLvl = 0
        #parLvl = 0
        #parKey = ""
        MoreColumns = True
             
        ercRet = self.GetMD(MDList0, udWBObj, self.sheetName, row, col, indentLvl, MoreColumns, "N")
        if (-1 == ercRet):
            print("GetMD failed")
            return
        #self.TraverseMD(MDList0, 1)
        #print("=========== END OF GETUD ===========")
        
        udWSObj = self.topUDObj.get_sheetObj(self.sheetName)
        if (None != udWSObj):
            pass
        else:
            return -1

        TCNameInd = 1
        self.GetUD(SyntaxList, 
                   MDList0, 
                   UDList0, 
                   udWBObj, 
                   self.sheetName, 
                   TCNameInd, 
                   self.udKey, 
                   -1, 
                   indentLvl,
                   self.udKey)
        udWBObj.close()
        print("--------------TRAVERSEUD--------------")
        self.TraverseUD(MDList0, UDList0, indentLvl)
        
        #self.mdWBObj.close()
    #-------------------------
        if (0):
            MDList = DataNode1D()
            UDList = DataNode1D()
        
            sheetName = "MDPar"
            sheetObj = self.topUDObj.get_sheetObj(sheetName)
            if (None != sheetObj):
                pass
            else:
                return -1
        
            ercRet = self.GetMD(MDList, sheetObj, sheetName, row, col, indentLvl, MoreColumns, "Y")
            if (-1 == ercRet):
                print("GetMD failed")
                return
                
            sheetComment = "#"
            udKey = "TSHSelenium"
            udkType = "M"
            keyMod = TSHKeyMod(sheetObj, self.THarnComment, self.MDKDelimiter, self.UDKDelimiter)
            DwareKeyAss = keyMod.GetUDKeys(udKey, udkType)
        
            nodeNameInd = 0
            nodeNameTot = 4
            nodeNameLst = ("TSHSelenium",
                           "TSHSelenium",
                           "c2",
                           "__init__")
            if (0):
                currNode = keyMod.TSHTreeNodesFind(keyMod.vKTree, keyMod.vKTree, nodeNameLst, nodeNameTot, nodeNameInd)
                if (None != currNode):
                    print(currNode.GetData())
                    for tNode in currNode.nodeArr:
                        print("  " + tNode.GetData())
                        print("  " + str(tNode.GetSSRowInd()))
                        SSRowInd = tNode.GetSSRowInd()
        
                        #self.TSHReadUDRec(sheetObj, MDList, UDList, SSRowInd)
                
                        MDColInd = 0
                        while(MDColInd < MDList.GetDataNode1DLen()):
                            MDNode = MDList.GetDataNode(MDColInd)
                            MwareDatAss = MDNode.MDAss
                            SSColInd = MwareDatAss["SSColInd"]
                            cellVal = self.sheetObj.cell(SSRowInd, SSColInd).value
                            MDColInd += 1
                        #kvDict["cbFuncKeyRoot"] = 
                        #kvDict["cbFuncId"] = 
                        #self.TSHFilterNode(kvDict)
            
    #------------------------------------------------------------------------------------------- 
    def TSHReadUDRec(self, MDList, UDList, SSRowInd):
        pass
    #------------------------------------------------------------------------------------------- 
    def TSHTreeNodesAddORIG(self, dataArr, ssLineNum):
        for dataInd in range(len(dataArr)):
            if (dataArr[dataInd] == None):
                break
        dataTot = dataInd
        
        currRootNode = self.root
        
        for dataInd in range(dataTot):
            dataEl = dataArr[dataInd]
            print(dataEl)            
            self.TSHTreeNodeAdd(currRootNode, dataEl, ssLineNum)                        


    

    #-------------------------------------------------------------------------------------------     
    def GetMDNew(self, DataList0, wbObj, sheetName, row, col, indentLvl, MoreColumns, ignoreRecursion):
        #print("GetMD:" + sheetName)            
        MwareDatAss = {}
        MwareKeyAss = {}
        
        sheetObj = self.topUDObj.get_sheetObj(sheetName)
        if (None != sheetObj):
            pass
        else:
            return -1

        keyMod2 = TSHKeyMod(sheetObj, self.MwareComment, self.MDKDelimiter, self.UDKDelimiter)
        MwareKeyAss = keyMod2.GetMDKeys(self.mdKey, self.mdkType)
        col = keyMod2.EOMDColInd + 1
        #for key in (MwareKeyAss):
        #    print("MwareKeyAss[" + str(key) + "]:" + str(MwareKeyAss[key]))
            #MwareDatAss[key] = sheetObj.cell(DataTypeCol, col).value
        
        
        
        AttributeNameAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "AttributeName" 
        DataTypeAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "DataType"
        TabAttr = self.mdKey + self.MDKDelimiter + "TSH" + self.MDKDelimiter + "Tab"
        
        
        AttributeNameCol = -1
        DataTypeCol = -1
        TabCol = -1
        for key in (MwareKeyAss):
            #print(str(key) + ":" + str(MwareKeyAss[key]))
            val = MwareKeyAss[key]
            
            x = val.rfind(AttributeNameAttr, 0, len(val))
            if (-1 != x):
                AttributeNameCol = int(key)

            x = val.rfind(DataTypeAttr, 0, len(val))
            if (-1 != x):
                DataTypeCol = int(key)
                    
            x = val.rfind(TabAttr, 0, len(val))
            if (-1 != x):
                TabCol = int(key)

        if (-1 == AttributeNameCol):
            print("Error: There is a problem with the MD Key Setup for " + str(self.mdKey))
            return -1
        
        MDNode0 = MDNode(MwareKeyAss)
        DataList0.AddDataNode(MDNode0)

        while (True == MoreColumns):

            commentCellValue = sheetObj.cell(1, col).value
            ercRet = self.MwareComment.isAComment(commentCellValue)
            ercRet = self.MwareComment.ProcessCommentStatus()
            if (ercRet == -1):
                return ercRet
            AttributeName = sheetObj.cell(AttributeNameCol, col).value
            if (ercRet == 1) and (AttributeName is None):
                col += 1
                continue
            elif (ercRet == 1) and (AttributeName is not None):
                col += 1
                continue
            elif (ercRet == 0) and (AttributeName is None):
                MoreColumns = False
            elif (ercRet == 0) and (AttributeName is not None):

                for  key in (MwareKeyAss):
                    #if (-1 != key):
                    MwareDatAss[key] = sheetObj.cell(int(key), col).value
                
                DataType = sheetObj.cell(DataTypeCol, col).value
                if (DataType is None): DataType = "None"
                
                Tab = sheetObj.cell(TabCol, col).value
                if (Tab is None): Tab = "None"
                #self.PrintMDCol(AttributeName, DataType, Tab)

                MwareDatAss["SSColInd"] = col
                MDNode0 = MDNode(MwareDatAss)
                
                if ("array" == DataType):
                    if ("Y" == ignoreRecursion):
                        DataList0.AddDataNode(MDNode0)
                    else:
                        MoreColumns = True
                        indentLvl += 1
                        DataList0.AddDataNode(MDNode0)
                        ercRet = self.GetMDNew(MDNode0.GetDataNode1D(), wbObj, Tab, 1, 1, indentLvl, MoreColumns, "")
                        if (-1 == ercRet):
                            return ercRet
                        indentLvl -= 1
                else:
                    DataList0.AddDataNode(MDNode0)
                col += 1
            else:
                return -1
        
        

        return 0
                
    #------------------------------------------------------------------------------------------- 
    
    def GetUDORIG(self, SyntaxList, MDList0, UDList0, wbObj, sheetName, TCNameInd, TCNamePrefix, UDRowInd, indentLvl):
        
        sheetObj = self.topUDObj.get_sheetObj(sheetName)
        if (None != sheetObj):
            pass
        else:
            return -1
        
        moreChildren = True

        
        while(True == moreChildren):
        # GETTING NEXT ROW FROM WS
            XLRowInd, XLColInd = self.GetEOMDRowAndCol(sheetObj)
            if ((-1 == XLRowInd)or(-1 == XLColInd)):
                break
            else:
                1#print (":colInd:"+str(XLColInd)+":rowInd:"+str(XLRowInd))
            
            ercRet = -2
            XLColInd = 3
            while (-2 == ercRet):
                XLRowInd += 1
                ercRet, TCName, TCNameInd = self.GetUDLocation(wbObj, sheetObj, XLRowInd, XLColInd, TCNameInd, TCNamePrefix)
                TCNameInd += 1
            TCNameInd -= 1
                
            XLRowInd = ercRet
            
            if (-1 == XLRowInd):
                moreChildren = False
                break
            #elif (-2 == XLRowInd):
            #    print ("SKIP PROCESSING...")
            #    sys.exit()
            else:
                #ADDING ROW OF DATA
                UD1D = DataNode1D()
                UDList0.AddDataNode1D(UD1D)
                UDRowInd += 1
                
                MDColInd = 0
                UDColInd = 0
                #sheetObj = wbObj[sheetName]
                
                dataVal = sheetObj.cell(XLRowInd, 2).value
                UDNode0 = DataNode(dataVal)
                UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0)
#                opStr = UDList0.GetDataNode1D(UDRowInd).GetDataNode(0).DispObjVal(self.GetPadStr(indentLvl), "TestCaseDelimiter")

              
#                dataVal = sheetObj.cell(XLRowInd, 2).value
#                print ("2:" + str(dataVal))
#                UDNode0 = DataNode(dataVal)
#                UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0)
#                opStr = UDList0.GetDataNode1D(UDRowInd).GetDataNode(0).DispObjVal(self.GetPadStr(indentLvl), "TestCaseDelimiter")
                #print(UDList0.GetDataNode1D(UDRowInd).GetDataNode(0).DispObjVal(self.GetPadStr(indentLvl), "TestCaseDelimiter"))
                
                #if (1 == self.debugId):
                dataVal = sheetObj.cell(XLRowInd, 3).value
                UDNode0 = DataNode(dataVal)
                UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0)
 #               opStr = UDList0.GetDataNode1D(UDRowInd).GetDataNode(1).DispObjVal(self.GetPadStr(indentLvl), "TestCase")
                    #print (opStr)
                #else:
                    #print (opStr)
                    #pass
                
                while(MDColInd < MDList0.GetDataNode1DLen()):
                    MDNode0 = MDList0.GetDataNode(MDColInd)
            
                    XLColInd = MDNode0.GetCol()
                    ScopeId = MDNode0.GetScopeId()
                    SyntaxId = MDNode0.GetSyntaxId()
                    dataVal = sheetObj.cell(XLRowInd, XLColInd).value
                    cell  = sheetObj.cell(XLRowInd, XLColInd)
                    if (cell is not None):
                        if (cell.comment is not None):
                            comment = str(cell.comment)
                            #xArr = comment.split()
                            xArr = comment.splitlines()
                            #xArr = comment.split('\n')
                            for xEl in (xArr):
                                xEl = xEl.replace("Comment: ", "", 1)
                                xEl = xEl.replace(" by Alexander Raj", "", 1)
                                #print ("Line:" + xEl)
                                xArr1 = xEl.split('=')
                                print (len(xArr1))
                                if (len(xArr1) == 2):
                                    xArr1[0] = xArr1[0].strip()
                                    xArr1[1] = xArr1[1].strip()
                                    #xArr1[0] = strip(xArr1[0])
                                    #xArr1[1] = strip(xArr1[1])
                                    print("key:", xArr1[0])
                                    print("val:", xArr1[1])
                            #print ("-------------comment:" + str(comment))
                    
    #                    print(sheetObj.cell(XLRowInd, XLColInd). .comment.text)
                    
                    #print (":WBS:"+sheetName+":TCP:"+TCNamePrefix+":TC:"+TCName+":MD:"+str(MDColInd)+":UD:"+str(UDRowInd)+":"+str(UDColInd)+":XL:"+str(XLRowInd) + ":" + str(XLColInd)+":DT:"+MDNode0.GetDataType()+":val:"+str(dataVal))        
                    #print (str(dataVal))
                    opStr = ""
                    if (SyntaxId is not None):
                        xRec = self.LookupSyntaxRec(SyntaxList, SyntaxId)
                        if (xRec is not None):
                            if (xRec.GetAction() is not None):
                                #opStr += xRec.GetAction() + " "
                                pass
                            if (xRec.GetResourceType() is not None):
                                #opStr += xRec.GetResourceType() + " "
                                pass
    
                    UDNode0 = DataNode(dataVal)
                    UDList0.GetDataNode1D(UDRowInd).AddDataNode(UDNode0) #Add after rowAttrs from header definition such as testCaseName, headerIndex?
                    opStr += UDList0.GetDataNode1D(UDRowInd).GetDataNode(UDColInd+2).DispObjVal(self.GetPadStr(indentLvl), MDNode0.GetAttributeName())
                    print ("opStr:" + opStr)
                    
                    if ("array" == MDNode0.GetDataType()):
                        if ("n" != dataVal):
                            #break
                        #print ("recur..." + MDNode0.GetTab())
                            ercRet = self.GetUD(SyntaxList, MDNode0.GetDataNode1D(), UDNode0.GetDataNode2D(), wbObj, MDNode0.GetTab(), 1, TCName, -1, indentLvl+1)
                        #print ("AFTER GETUD...")
                    MDColInd += 1
                    UDColInd += 1
                    #print ("end of while2")
            TCNameInd += 1
            #print ("end of while1")
        return 0

    #------------------------------------------------------------------------------------------- 
    