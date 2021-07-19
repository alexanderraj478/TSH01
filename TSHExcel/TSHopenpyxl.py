import openpyxl as tshwb
from openpyxl import Workbook
from TSHGlobal import constants
from TSHFS.TSHFile import TSHFile
from TSHConvert.TSHODTToString import TSHString
from TSHPrint.TSHPrintStr import TSHPrintStr

class TSHopenpyxl:
    def __init__(self, filePath):
        self.wbObj = None
        self.wsObj = None
        self.vtshFile = TSHFile()
        self.vtshFile.filePath = filePath
        self.xlsRow = 1
        self.xlsCol = 0

    def WSAdd(self, sheetName):
        self.wsObj = self.wbObj.create_sheet(sheetName)

    def WSCellWrite(self, row, col, cellVal):
        #self.wsObj.write(row, col, cellVal)
        if (None == cellVal):
            cellVal = ""
        self.wsObj.cell(row=row, column=col).value = cellVal

    def WBSave(self):
        self.wbObj.save(self.vtshFile.filePath)  #'/home/alex/test/TSHTestware.xls'
        
    def WBClose(self):
        if hasattr(self, "wbObj"):
            if (None != self.wbObj):
                #print(self.wbObj.__dict__)
                #print(vars(self.wbObj))
#                print(tshwb.__version__)
#                self.wbObj.close()
                #if (hasattr(self.wbObj, ))
                1 #??? this seems to fail self.wbObj.close()

    def WBSaveAndClose(self):
        self.wbObj.save(self.vtshFile.filePath)  #'/home/alex/test/TSHTestware.xls'
        self.WBClose()
                
    def WBCreate(self):
        self.wbObj = Workbook()
        
    def WBGetObj(self):
        return self.wbObj
        
    def WSGetObj(self):
        return self.wsObj

    def Initialize(self, file_path):
        #is this good design?
        #self.vtshFile.FilePathSet(file_path)
        #self.vtshFile. .Initialize(file_path)
        pass
        
    def WBOpen(self):
        if (True == self.vtshFile.TSHPathExists()):
            self.wbObj = tshwb.load_workbook(self.vtshFile.filePath)
            return self.wbObj
        else:
            return None

    def WBObjClose(self, tshwb):
        if (self.wbObj != None):
            tshwb.close(tshwb)
        else:
            print("Invalid workbook object passed")
            
    def workbookActiveSheetSet(self, sheetName):
        if (sheetName in self.get_sheetnames()):
            self.wbObj.active = self.wbObj[sheetName]      
            return self.wsObj
        else:
            vtshPrtMsg = TSHPrintStr("Error: {}:{}: {}".format(-2, constants.TSHFailDict["-2"], sheetName))
            vtshPrtMsg.PrintMsg3 ()
            return None

    def get_sheetObj(self, sheetName):
        if (sheetName in self.WBObjGetSheetNames()):
            return self.wbObj[sheetName]
        else:
            #constants.PrintMsg3 ("Error: {}:{}: {}".format(-2, constants.TSHFailDict["-2"], sheetName))
            return None
            
    def WBObjDelWorksheet1(self, sheetName):
        del self.wbObj.defined_names[sheetName]
    
    def WBObjDelWorksheet2(self, sheetName):
        self.wbObj.remove(sheetName)
    
    def WSObjGetSheetName(self, sheetObj):
        return sheetObj.title
    
    def WBObjGetSheetNames(self):
        return self.wbObj.sheetnames
        
    def WBObjGetWorksheets(self):
        return self.wbObj.worksheets
    